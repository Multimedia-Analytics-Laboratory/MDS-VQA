import json
from scipy.stats import spearmanr, pearsonr
from openpyxl import load_workbook
from utils import fit_curve, contains_nan


def gMAD_srcc(gt_path, pred_path, dataset, sheet, min_row=2, score_col=1, filter_list_path=None, start_idx=0, interval=None):
    mos = {}
    wb = load_workbook(gt_path, read_only=True)        
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        id = str(row[0]).strip()
        score = float(row[score_col])
        if id and score is not None:
            mos[id] = score
    wb.close()

    with open(pred_path) as f:
        pred_ori = json.load(f)
    
    if filter_list_path:
        filter_list = []
        if filter_list_path.endswith("txt"):
            with open(filter_list_path, "r") as f:
                for line in f.readlines():
                    line = line.strip('\n')
                    filter_list.append(line.split(" ")[0])
        else:
            with open(filter_list_path, "r") as f:
                filter_list = json.load(f)
                if interval:
                    filter_list = filter_list[start_idx:start_idx+interval]

    pred_list = []
    mos_list = []
    for item in pred_ori:
        if 'uvq_dataset' in gt_path:
            mos_item = "_".join(item.split("_")[:2])
        elif '.mp4' in item and item[:-4] in mos:
            mos_item = item[:-4]
        else:
            mos_item = item

        if mos_item not in mos:
            continue
        if filter_list_path and item not in filter_list:
            continue

        pred_list.append(pred_ori[item]["score"])
        mos_list.append(mos[mos_item])

    fitted_x = fit_curve(pred_list, mos_list)

    if contains_nan(fitted_x) or all(x == fitted_x[0] for x in fitted_x):
        fitted_x = pred_list

    srcc1 = spearmanr(fitted_x, mos_list)[0]
    plcc1 = pearsonr(fitted_x, mos_list)[0]

    print(f"Dataset size: {len(fitted_x)} ======================{start_idx}")
    print(f"{dataset} ori srcc: {srcc1}, plcc: {plcc1}")

    return srcc1, plcc1


if __name__ == '__main__':
    srcc_sdr, plcc_sdr = gMAD_srcc(
        'datasets/sdr/mos.xlsx',
        'test_results/sdr_active_finetuning.json',
        'YouTube-SFV SDR',
        'SDR',
        filter_list_path="datasets/sdr/sdr_test_set.txt"
    )
